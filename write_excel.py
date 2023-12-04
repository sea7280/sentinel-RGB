import openpyxl
import os
from osgeo import gdal
import sys
sys.dont_write_bytecode = True
from tkinter import filedialog
from tkinter import messagebox

import parameter
import path


#NDVI, FDIのエクセル書き込み
def saveExcel(entry):
    def write_excel(band,sheet):
        ws = sheet
        size_y = len(band)
        size_x = len(band[0])
        for i in range(size_y):
            for j in range(size_x):
                ws.cell(row=i + 1,column=j + 1).value = band[i][j]

    filepath = path.pathGet(entry)
    setting_detail = parameter.getParameters(entry)

    minX          = setting_detail[2]
    minY          = setting_detail[3]
    deltaX        = setting_detail[4]
    deltaY        = setting_detail[5]

    bluepath        = filepath[0]
    greenpath       = filepath[1]
    redpath         = filepath[2]
    nirpath         = filepath[3]

    band2_8bit_path =os.path.dirname(__file__) + "/Band2_8bit.tif"
    band3_8bit_path =os.path.dirname(__file__) + "/Band3_8bit.tif"
    band4_8bit_path =os.path.dirname(__file__) + "/Band4_8bit.tif"
    band8_8bit_path =os.path.dirname(__file__) + "/Band8_8bit.tif"

    gdal.Translate(band2_8bit_path ,bluepath    , srcWin=[minX,minY,deltaX,deltaY])
    gdal.Translate(band3_8bit_path ,greenpath   , srcWin=[minX,minY,deltaX,deltaY])
    gdal.Translate(band4_8bit_path ,redpath     , srcWin=[minX,minY,deltaX,deltaY])
    gdal.Translate(band8_8bit_path ,nirpath     , srcWin=[minX,minY,deltaX,deltaY])

    b2_image =gdal.Open(band2_8bit_path)
    b3_image =gdal.Open(band3_8bit_path)
    b4_image =gdal.Open(band4_8bit_path)
    b8_image =gdal.Open(band8_8bit_path)

    BlueBand_array      = b2_image.ReadAsArray()
    GreenBand_array     = b3_image.ReadAsArray()
    RedBand_array       = b4_image.ReadAsArray()
    NIR_Band_array      = b8_image.ReadAsArray()

    BlueBand_array     = BlueBand_array     / 10000
    GreenBand_array    = GreenBand_array    / 10000
    RedBand_array      = RedBand_array      / 10000
    NIR_Band_array     = NIR_Band_array     / 10000

    print(BlueBand_array)
    print(GreenBand_array)
    print(RedBand_array)

################################################ save ##########################################################

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = "Blue"
    ws = wb.create_sheet(title="Green")
    ws = wb.create_sheet(title="Red")
    ws = wb.create_sheet(title="NIR")


    write_excel(BlueBand_array        ,wb["Blue"])
    write_excel(GreenBand_array       ,wb["Green"])
    write_excel(RedBand_array         ,wb["Red"])
    write_excel(NIR_Band_array        ,wb["NIR"])


    ftype=[('Excelファイル', '*.xlsx')]
    idir=os.path.expanduser('~/Desktop')
    savefile=filedialog.asksaveasfilename(defaultextension='xlsx', filetypes=ftype, initialdir=idir)
    if savefile!='':
        wb.save(savefile)
    
    os.remove(band2_8bit_path)
    os.remove(band3_8bit_path)
    os.remove(band4_8bit_path)
    os.remove(band8_8bit_path)
    if savefile!='':
        messagebox.showinfo("", "All complete.")
